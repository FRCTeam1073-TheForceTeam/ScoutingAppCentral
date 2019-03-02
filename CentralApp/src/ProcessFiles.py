'''
Created on Feb 4, 2014

@author: ken_sthilaire
'''
import os
import re
import traceback
import time
import datetime
import sys
import openpyxl
import shutil
import json

from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.cell import get_column_letter
from copy import copy

from optparse import OptionParser

import AttributeDefinitions
import ConfigUtils
import CompAlias
import DbSession
import DataModel
import IssueTrackerDataModel
import DebriefDataModel
import FileParser
import Logger
import WebCommonUtils
import FileSync


def log_exception(logger, e):
    logger.debug('Exception Caught Processing Files: %s' % str(e) )
    traceback.print_exc(file=sys.stdout)
    exc_type, exc_value, exc_traceback = sys.exc_info()
    exception_info = traceback.format_exception(exc_type, exc_value,exc_traceback)
    for line in exception_info:
        line = line.replace('\n','')
        logger.debug(line)


'''Get list of files to be processed.

Args:
    input_dir: Directory to search.
    pattern: Regular expression to use to filter files.
    recursive: Whether or not to recurse into input_dir.

Returns:
    A list of files.
'''
def isFileProcessed(global_config, session, db_name, filepath):
    if db_name == (global_config['db_name'] + global_config['this_season']):
        is_processed = DataModel.isFileProcessed(session, filepath)
    elif db_name == (global_config['issues_db_name'] + global_config['this_season']):
        is_processed = IssueTrackerDataModel.isFileProcessed(session, filepath)
    elif db_name == (global_config['debriefs_db_name'] + global_config['this_season']):
        is_processed = DebriefDataModel.isFileProcessed(session, filepath)
        
    return is_processed
                                                 
def get_files(global_config, session, db_name, input_dir, pattern, recursive=True):    
    file_list = []    
        
    if recursive:
        for root, dirs, files in os.walk(input_dir):

            #print 'Root:', root, ' Dirs: ', dirs, ' Files:', files
            for name in files:
                if pattern.match(name):
                    pathname = os.path.join(root, name)
                    file_processed = isFileProcessed(global_config, session, db_name, pathname)
                    if file_processed is False:
                        if pathname not in file_list:
                            global_config['logger'].debug( '%s - Adding %s to be processed' % (__name__,pathname))
                            file_list.append(pathname)
                        else:
                            global_config['logger'].debug( '%s - File %s already in list' % (__name__,pathname))
    else:
        files = os.listdir(input_dir)
        #print 'Files:', files
        for name in files:
            if pattern.match(name):
                pathname = os.path.join(input_dir, name)
                file_processed = isFileProcessed(global_config, session, db_name, pathname)
                if file_processed is False:
                    if pathname not in file_list:
                        global_config['logger'].debug( '%s - Adding %s to be processed' % (__name__,pathname))
                        file_list.append(pathname)
                    else:
                        global_config['logger'].debug( '%s - File %s already in list' % (__name__,pathname))

    if len(file_list) > 0:
        print 'FileList:', file_list
        
    return file_list

  
def process_files(global_config, attr_definitions, input_dir, recursive=True):
    start_time = datetime.datetime.now()

    # Initialize the database session connection
    db_name  = global_config['db_name'] + global_config['this_season']
    session  = DbSession.open_db_session(db_name)
 
    some_files_processed = False
    
    # read the ignore file list config each time through the loop. Any files
    # in the ignore list will be skipped
    ignore_filelist = read_ignore_filelist_cfg(input_dir + 'IgnoreFiles.txt')

    # The following regular expression will select all files that conform to 
    # the file naming format Team*.txt. Build a list of all datafiles that match
    # the naming format within the directory passed in via command line 
    # arguments.
    file_regex = re.compile('Team[a-zA-Z0-9_]+.txt')
    files = get_files(global_config, session, db_name, input_dir, file_regex, recursive)
    
    if len(files) > 0:
        log_msg = 'files retrieved, elapsed time - %s' % (str(datetime.datetime.now()-start_time))
        print log_msg
        global_config['logger'].debug( '%s - %s' % (process_files.__name__,log_msg))

        global_config['logger'].debug( '%s - %d Files to be processed' % (process_files.__name__,len(files)))
        
    # Process data files
    for data_filename in files:
        # If the file is on the ignore list (quarantined), then skip it
        if data_filename.split('/')[-1] in ignore_filelist:
            global_config['logger'].debug( '%s - Ignoring file: %s' % (process_files.__name__,data_filename))
            continue
        
        # Make sure that the data file has not already been processed. We have seen cases
        # where the data file gets inserted into the list of files to be processed more than
        # once.
        file_processed = isFileProcessed(global_config, session, db_name, data_filename)
        if not file_processed:
            try:
                global_config['logger'].debug( '%s - Processing file: %s' % (process_files.__name__,data_filename))
                process_file( global_config, session, attr_definitions, data_filename)
            except Exception, e:
                global_config['logger'].debug( '%s - Error processing file: %s' % (process_files.__name__,data_filename))
                # log the exception but continue processing other files
                log_exception(global_config['logger'], e)

            # add the file to the set of processed files so that we don't process it again. Do it outside the
            # try/except block so that we don't try to process a bogus file over and over again.       
            DataModel.addProcessedFile(session, data_filename)
            some_files_processed = True
        else:
            global_config['logger'].debug( '%s - Skipping file: %s, already processed' % (process_files.__name__,data_filename))
            
        # Commit all updates to the database
        session.commit()
        
    if some_files_processed == True:    
        log_msg = 'files processed, elapsed time - %s' % (str(datetime.datetime.now()-start_time))
        print log_msg
        global_config['logger'].debug( '%s - %s' % (process_files.__name__,log_msg))
    
        DataModel.dump_database_as_csv_file(session, global_config, attr_definitions)
    
        log_msg = 'database dumped, elapsed time - %s' % (str(datetime.datetime.now()-start_time))
        print log_msg
        global_config['logger'].debug( '%s - %s' % (process_files.__name__,log_msg))
            
    session.remove()
        
def process_file(global_config, session, attr_definitions, data_filename):
    print 'processing %s'%data_filename
    
    # Initialize the file_attributes dictionary in preparation for the
    # parsing of the data file
    file_attributes = {}
    
    # Parse the data file, storing all the information in the file_attributes
    # dictionary
    FileParser.FileParser(data_filename).parse(file_attributes)

    # The team number can be retrieved from the Team attribute, one of the
    # mandatory attributes within the data file
    team = file_attributes['Team']
    
    # Also, extract the competition name, too, if it has been included in
    # the data file
    if file_attributes.has_key('Competition'):
        # check if the global_config indicates that we're working with the 2014-era tablet UI that may not
        # format the competition string as we expect it to be. If we are in 'legacy' mode, then ignore the
        # competition setting in the file and apply the competition/season from the config
        if global_config.has_key('legacy_tablet_ui') and global_config['legacy_tablet_ui'].lower() == 'yes':
            competition = global_config['this_competition'] + global_config['this_season']
        else:
            competition = file_attributes['Competition']
    else:
        # if no competition setting attribute is in the file, then apply the competition/season from the config
        competition = global_config['this_competition'] + global_config['this_season']

        if competition == None:
            raise Exception( 'Competition Not Specified!')

    DataModel.addTeamToEvent(session, team, competition)
    
    if file_attributes.has_key('Scouter'):
        scouter = file_attributes['Scouter']
    else:
        scouter = 'Unknown'
        
    if '_Match' in data_filename:
        category = 'Match'
        if not file_attributes.has_key('Match'):
            file_attributes['Match'] = '0'
    elif '_Pit_' in data_filename:
        category = 'Pit'
    else:
        category = 'Other'

    # Loop through the attributes from the data file and post them to the
    # database
    for attribute, value in file_attributes.iteritems():
        if value is None:
            value = ''
        try:
            attr_definition = attr_definitions.get_definition(attribute)
            if attr_definition == None:
                err_str = 'ERROR: No Attribute Defined For Attribute: %s' % attribute
                print err_str
            elif attr_definition['Database_Store']=='Yes':
                try:
                    DataModel.createOrUpdateAttribute(session, team, competition, category, attribute, value, attr_definition)
                except Exception, exception:
                    traceback.print_exc(file=sys.stdout)
                    exc_type, exc_value, exc_traceback = sys.exc_info()
                    exception_info = traceback.format_exception(exc_type, exc_value,exc_traceback)
                    for line in exception_info:
                        line = line.replace('\n','')
                        global_config['logger'].debug(line)
        except Exception:
            err_str = 'ERROR: Attribute Could Not Be Processed: %s' % attribute
            traceback.print_exc(file=sys.stdout)
            exc_type, exc_value, exc_traceback = sys.exc_info()
            exception_info = traceback.format_exception(exc_type, exc_value,exc_traceback)
            for line in exception_info:
                line = line.replace('\n','')
                global_config['logger'].debug(line)
            print err_str
            
        if category == 'Match':
            try:
                match = file_attributes['Match']
                DataModel.createOrUpdateMatchDataAttribute(session, team, competition, match, scouter, attribute, value)
            except:
                err_str = 'ERROR: Match Data Attribute Could Not Be Processed: %s' % attribute
                traceback.print_exc(file=sys.stdout)
                exc_type, exc_value, exc_traceback = sys.exc_info()
                exception_info = traceback.format_exception(exc_type, exc_value,exc_traceback)
                for line in exception_info:
                    line = line.replace('\n','')
                    global_config['logger'].debug(line)
                print err_str
            
    score = DataModel.calculateTeamScore(session, team, competition, attr_definitions)
    DataModel.setTeamScore(session, team, competition, score)

        
def remove_file_data(global_config, session, attr_definitions, data_filename, remove_from_processed_files=False):
    print 'removing data file: %s'%data_filename
    
    competition = global_config['this_competition'] + global_config['this_season']
    input_dir = './static/data/' + competition + '/ScoutingData/'
    filepath = input_dir+data_filename
    # Initialize the file_attributes dictionary in preparation for the
    # parsing of the data file
    file_attributes = {}
    
    # Parse the data file, removing all the information in the file_attributes
    # dictionary
    try:
        FileParser.FileParser(filepath).parse(file_attributes)
    except:
        raise ValueError('Error Opening File')

    # The team number can be retrieved from the Team attribute, one of the
    # mandatory attributes within the data file
    team = file_attributes['Team']
    
    # Also, extract the competition name, too, if it has been included in
    # the data file
    if file_attributes.has_key('Competition'):
        competition = file_attributes['Competition']
    else:
        competition = global_config['this_competition'] + global_config['this_season']

        if competition == None:
            raise Exception( 'Competition Not Specified!')
    
    # Loop through the attributes from the data file and post them to the
    # database
    for attribute, value in file_attributes.iteritems():
        if value is None:
            value = ''
        attr_definition = attr_definitions.get_definition(attribute)
        if attr_definition == None:
            raise ValueError( 'No Attribute Defined For Attribute: %s' % attribute )
        elif attr_definition['Database_Store']=='Yes':
            DataModel.deleteAttributeValue(session, team, competition, attribute, value, attr_definition, no_throw=True)
            
    if remove_from_processed_files:
        DataModel.deleteProcessedFile(session, filepath)

    score = DataModel.calculateTeamScore(session, team, competition, attr_definitions)
    DataModel.setTeamScore(session, team, competition, score)

def process_issue_files(global_config, input_dir, recursive=True):

    # Initialize the database session connection
    issues_db_name  = global_config['issues_db_name'] + global_config['this_season']
    debrief_db_name = global_config['debriefs_db_name'] + global_config['this_season']
    debrief_session = DbSession.open_db_session(debrief_db_name)
    issues_session  = DbSession.open_db_session(issues_db_name)

    # The following regular expression will select all files that conform to 
    # the file naming format Issue*.txt. Build a list of all datafiles that match
    # the naming format within the directory passed in via command line 
    # arguments.
    file_regex = re.compile('Issue[a-zA-Z0-9_-]+.txt')
    files = get_files(global_config, issues_session, issues_db_name, input_dir, file_regex, recursive)
    
    files.sort()
    
    # Process data files
    for data_filename in files:
        print 'processing %s'%data_filename
        try:
            # Initialize the file_attributes dictionary in preparation for the
            # parsing of the data file
            issue_attributes = {}
            
            # Parse the data file, storing all the information in the file_attributes
            # dictionary
            FileParser.FileParser(data_filename).parse(issue_attributes)
    
            issue = IssueTrackerDataModel.addIssueFromAttributes(issues_session, issue_attributes)
            if issue.debrief_key != None:
                match_str, issue_key = issue.debrief_key.split('_')
                competition = global_config['this_competition'] + global_config['this_season']

                DebriefDataModel.addOrUpdateDebriefIssue(debrief_session, int(match_str), 
                                                         competition,
                                                         issue.issue_id, issue_key)
        except Exception, e:
            # log the exception but continue processing other files
            log_exception(global_config['logger'], e)

        # add the file to the set of processed files so that we don't process it again. Do it outside the
        # try/except block so that we don't try to process a bogus file over and over again.       
        IssueTrackerDataModel.addProcessedFile(issues_session, data_filename)
            
    debrief_session.commit()
    issues_session.commit()
    debrief_session.remove()
    issues_session.remove()
    
    
def process_debrief_files(global_config, input_dir, recursive=True):

    # Initialize the database session connections
    issues_db_name  = global_config['issues_db_name'] + global_config['this_season']
    debrief_db_name = global_config['debriefs_db_name'] + global_config['this_season']
    debrief_session = DbSession.open_db_session(debrief_db_name)
    issues_session  = DbSession.open_db_session(issues_db_name)
    
    # Create the database if it doesn't already exist
    #if not os.path.exists('./' + db_name):    
    #   DebriefDataModel.create_db_tables(my_db)

    # The following regular expression will select all files that conform to 
    # the file naming format Debrief*.txt. Build a list of all datafiles that match
    # the naming format within the directory passed in via command line 
    # arguments.
    file_regex = re.compile('Debrief[a-zA-Z0-9_-]+.txt')
    files = get_files(global_config, debrief_session, debrief_db_name, input_dir, file_regex, recursive)
    
    # Process data files
    for data_filename in files:
        print 'processing %s'%data_filename
        try:
            # Initialize the debrief_attributes dictionary in preparation for the
            # parsing of the data file
            debrief_attributes = {}
            
            # Parse the data file, storing all the information in the attributes
            # dictionary
            FileParser.FileParser(data_filename).parse(debrief_attributes)
            DebriefDataModel.addDebriefFromAttributes(debrief_session, debrief_attributes)
            
            # Also, extract the competition name, too, if it has been included in
            # the data file
            
            if debrief_attributes.has_key('Competition'):
                competition = debrief_attributes['Competition']
                issue_base_name = WebCommonUtils.split_comp_str(competition)[0]
            else:
                competition = global_config['this_competition'] + global_config['this_season']
                issue_base_name = global_config['this_competition']

                if competition == None:
                    raise Exception( 'Competition Not Specified!')
    
            # At competition, we will likely have multiple laptops manging the data, but we want
            # only one machine to be responsible for the issues database. In all likelihood,
            # that machine will be the one in the pits, or possibly the application running
            # in the cloud.
            if global_config['issues_db_master'] == 'Yes':
                match_id = debrief_attributes['Match']
                submitter = debrief_attributes['Scouter']
                timestamp = str(int(time.time()))
                subgroup = 'Unassigned'
                status = 'Open'
                owner = 'Unassigned'
                
                if debrief_attributes.has_key('Issue1_Summary') or debrief_attributes.has_key('Issue1_Description'):
                    # look to see if there is already a debrief issue, and if so, do not attempt to create/update
                    # an issue, as there are already other issue files that would then conflict with this one
                    issue_key = 'Issue1'
                    if DebriefDataModel.getDebriefIssue(debrief_session, competition, match_id, issue_key) == None:
                        # if no summary is provided, then use the description as the summary. Likewise, if no description
                        # is provided then use the summary as the description. Keep in mind that we need at least the
                        # summary or description to be provided.
                        if debrief_attributes.has_key('Issue1_Summary'):
                            summary = debrief_attributes['Issue1_Summary']
                        else:
                            summary = debrief_attributes['Issue1_Description']
                        if debrief_attributes.has_key('Issue1_Description'):
                            description = debrief_attributes['Issue1_Description']
                        else:
                            description = debrief_attributes['Issue1_Summary']
                            
                        if debrief_attributes.has_key('Issue1_Priority'):
                            priority = debrief_attributes['Issue1_Priority']
                        else:
                            priority = 'Priority_3'
                            
                        if debrief_attributes.has_key('Issue1_Taskgroup'):
                            component = debrief_attributes['Issue1_Taskgroup']
                        else:
                            component = ''
                            
                        debrief_key = str(match_id) + '_' + issue_key
                    
                        issue_id = IssueTrackerDataModel.getIssueId(issues_session, issue_base_name)
                        issue = IssueTrackerDataModel.addOrUpdateIssue(issues_session, issue_id, summary, status, priority, 
                                 subgroup, component, submitter, owner, description, timestamp, debrief_key)
                        if issue != None:
                            issue.create_file('./static/data/%s/ScoutingData' % competition)
                        DebriefDataModel.addOrUpdateDebriefIssue(debrief_session, match_id, competition, issue_id, issue_key)
                    
                if debrief_attributes.has_key('Issue2_Summary') or debrief_attributes.has_key('Issue2_Description'):
                    # look to see if there is already a debrief issue, and if so, do not attempt to create/update
                    # an issue, as there are already other issue files that would then conflict with this one
                    issue_key = 'Issue2'
                    if DebriefDataModel.getDebriefIssue(debrief_session, competition, match_id, issue_key) == None:
                        # if no summary is provided, then use the description as the summary. Likewise, if no description
                        # is provided then use the summary as the description. Keep in mind that we need at least the
                        # summary or description to be provided.
                        if debrief_attributes.has_key('Issue2_Summary'):
                            summary = debrief_attributes['Issue2_Summary']
                        else:
                            summary = debrief_attributes['Issue2_Description']
                        if debrief_attributes.has_key('Issue2_Description'):
                            description = debrief_attributes['Issue2_Description']
                        else:
                            description = debrief_attributes['Issue2_Summary']
                            
                        if debrief_attributes.has_key('Issue2_Priority'):
                            priority = debrief_attributes['Issue2_Priority']
                        else:
                            priority = 'Priority_3'
                            
                        if debrief_attributes.has_key('Issue2_Taskgroup'):
                            component = debrief_attributes['Issue2_Taskgroup']
                        else:
                            component = ''

                        debrief_key = str(match_id) + '_' + issue_key
                        
                        issue_id = IssueTrackerDataModel.getIssueId(issues_session, issue_base_name)
                        issue = IssueTrackerDataModel.addOrUpdateIssue(issues_session, issue_id, summary, status, priority, 
                                 subgroup, component, submitter, owner, description, timestamp, debrief_key)
                        if issue != None:
                            issue.create_file('./static/data/%s/ScoutingData' % competition)
                        DebriefDataModel.addOrUpdateDebriefIssue(debrief_session, match_id, competition, issue_id, issue_key)
                    
                if debrief_attributes.has_key('Issue3_Summary') or debrief_attributes.has_key('Issue3_Description'):
                    # look to see if there is already a debrief issue, and if so, do not attempt to create/update
                    # an issue, as there are already other issue files that would then conflict with this one
                    issue_key = 'Issue3'
                    if DebriefDataModel.getDebriefIssue(debrief_session, competition, match_id, issue_key) == None:
                        # if no summary is provided, then use the description as the summary. Likewise, if no description
                        # is provided then use the summary as the description. Keep in mind that we need at least the
                        # summary or description to be provided.
                        if debrief_attributes.has_key('Issue3_Summary'):
                            summary = debrief_attributes['Issue3_Summary']
                        else:
                            summary = debrief_attributes['Issue3_Description']
                        if debrief_attributes.has_key('Issue3_Description'):
                            description = debrief_attributes['Issue3_Description']
                        else:
                            description = debrief_attributes['Issue3_Summary']
                            
                        if debrief_attributes.has_key('Issue3_Priority'):
                            priority = debrief_attributes['Issue3_Priority']
                        else:
                            priority = 'Priority_3'
                            
                        if debrief_attributes.has_key('Issue3_Taskgroup'):
                            component = debrief_attributes['Issue3_Taskgroup']
                        else:
                            component = ''

                        debrief_key = str(match_id) + '_' + issue_key
                    
                        issue_id = IssueTrackerDataModel.getIssueId(issues_session, issue_base_name)
                        issue = IssueTrackerDataModel.addOrUpdateIssue(issues_session, issue_id, summary, status, priority, 
                                 subgroup, component, submitter, owner, description, timestamp, debrief_key)
                        if issue != None:
                            issue.create_file('./static/data/%s/ScoutingData' % competition)
                        DebriefDataModel.addOrUpdateDebriefIssue(debrief_session, match_id, competition, issue_id, issue_key)
        except Exception, e:
            # log the exception but continue processing other files
            log_exception(global_config['logger'], e)

        # add the file to the set of processed files so that we don't process it again. Do it outside the
        # try/except block so that we don't try to process a bogus file over and over again.       
        DebriefDataModel.addProcessedFile(debrief_session, data_filename)
    
    issues_session.commit()
    debrief_session.commit()
    debrief_session.remove()
    issues_session.remove()

def read_ignore_filelist_cfg(ignore_filename):
    ignore_filelist = []
    if os.path.exists(ignore_filename):
        try:
            cfg_file = open(ignore_filename, 'r')
            for cfg_line in cfg_file:
                if cfg_line.startswith('#'):
                    continue
                cfg_items = cfg_line.split(',')
                for cfg_item in cfg_items:
                    ignore_filelist.append(cfg_item.rstrip())
            cfg_file.close()
        except:
            print 'Error reading ignore file configuration'
            
    return ignore_filelist

def match_sort(key):
    match_prefix = key.split('_')[0]
    try:
        match_number = int(match_prefix.replace('Match',''))
    except:
        match_number = 0
    return match_number

def process_json_files(global_config, competition, output_file, input_dir, reprocess_files=False):
    
    # Initialize the database session connection
    db_name  = global_config['db_name'] + global_config['this_season']
    session  = DbSession.open_db_session(db_name)

    # get all the verified files from the input directory. These files are
    # candidates to be processed
    verified_files = FileSync.get_file_list(input_dir, ext='.verified', recurse=True )
    verified_files.sort(key=match_sort)
    
    # For the normal case, get all the processed files, too. We'll use the processed list to 
    # determine which files are actually newly verified and need to be processed. If the 
    # reprocess flag is true, then we'll process all verified files.
    if reprocess_files is not True:
        processed_files = FileSync.get_file_list(input_dir, ext='.processed', recurse=True )
        for processed_file in processed_files:
            verified_files.remove( processed_file.replace('processed','verified') )
    
    # read in the output file, which is expected to be an XLSX file
    xlsx_workbook = openpyxl.load_workbook(output_file)
    
    '''
    # took out for now until we have local dictionary storage
    events = global_config.get('events')
    if events is None:
        events = {}
        global_config['events'] = events
    event_data = events.get( competition )
    if event_data is None:
        events[competition] = { 'ScoutingData': { 'TeamData': {} } }
        event_data = events[competition]

    event_scouting_data = event_data['ScoutingData']['TeamData']
    '''

    for verified_file in verified_files:
        # read the file into a dictionary
        with open(input_dir+verified_file) as fd:
            scouting_data = json.load(fd)
            team = scouting_data['Setup'].get('Team')

            if team is not None and len(team) > 0:
                # ######################################################### #
                # store the scouting data to the local database

                DataModel.addTeamToEvent(session, int(team), competition)

                attr_definitions = AttributeDefinitions.AttrDefinitions(global_config)
                for section_name, section_data in scouting_data.iteritems():
                    if isinstance(section_data,dict):
                        for attr_name, attr_value in section_data.iteritems():
                            
                            # use the attribute definitions to control whether information gets 
                            # stored to the database rather than the hard coded stuff here.
                            # also need to consider the section/category name as the attributes
                            # and definitions are processed
                            
                            # don't store the team number in the database
                            if attr_name == 'Team':
                                continue
                            
                            # augment the attribute name with the section name in order to make the attribute
                            # unique
                            attr_name = '%s:%s' % (section_name, attr_name)
                            
                            attribute_def = {}
                            attribute_def['Name'] = attr_name
                            if attr_value.isdigit():
                                attribute_def['Type'] = 'Integer'
                                attribute_def['Weight'] = 1.0
                            else:
                                attribute_def['Type'] = 'String'
                                attribute_def['Weight'] = 0.0
                            attribute_def['Statistic_Type'] = 'Average'
                            attr_definitions.add_definition(attribute_def)
                        
                            category = 'Match'
                            try:
                                DataModel.createOrUpdateAttribute(session, int(team), competition, category, 
                                                                  attr_name, attr_value, attribute_def)
                            except Exception, exception:
                                traceback.print_exc(file=sys.stdout)
                                exc_type, exc_value, exc_traceback = sys.exc_info()
                                exception_info = traceback.format_exception(exc_type, exc_value,exc_traceback)
                                for line in exception_info:
                                    line = line.replace('\n','')
                                    global_config['logger'].debug(line)
                    else:
                        print 'Unexpected entry in scouting data file, name: %s, value: %s' % (section_name,section_data)
                
                score = DataModel.calculateTeamScore(session, int(team), competition, attr_definitions)
                DataModel.setTeamScore(session, int(team), competition, score)
                session.commit()

                # ######################################################### #
                '''
                # ######################################################### #
                # store the scouting data to a local dictionary
                team_data = event_scouting_data.get(team)
                if team_data is None:
                    team_data = { 'Summary': {}, 'MatchData': [] }
                    event_scouting_data[team] = team_data

                team_match_data = team_data['MatchData']

                # if this match has already been scored, then update the data by removing the existing
                # match data and then add the updated data
                update_match = False
                for match_data in team_match_data:
                    if scouting_data['Setup']['Match'] == match_data['Setup']['Match']:
                        update_match = True
                        break
                if update_match is True:
                    team_match_data.remove(match_data)
                team_match_data.append(scouting_data)
                # ######################################################### #
                '''

                # ######################################################### #
                # store the scouting data information to the spreadsheet
                team_name = 'Team %s' % team
                try:
                    team_sheet = xlsx_workbook.get_sheet_by_name(team_name)
                except:
                    team_sheet = create_team_sheet( xlsx_workbook, team_name )

                curr_matches = team_sheet['B2'].value
                
                # get max row and column count and iterate over the sheet
                max_row= team_sheet.max_row
                
                for i in range(1,max_row+1):
                     # scan for a row that has Match in the first column to identify rows where data will be stored
                     cell_value = team_sheet.cell(row=i,column=1).value
                     if team_sheet.cell(row=i,column=1).value == 'Match':
                         attr_row = i
                         data_row = i+1
                         data_cell = team_sheet.cell(row=i+1,column=1).value
                         if data_cell is None:
                             team_sheet = update_data_row( team_sheet, attr_row, data_row, scouting_data )
                             team_sheet['B2'].value = curr_matches+1
                             shutil.copyfile(input_dir+verified_file, input_dir+verified_file.replace('verified','processed'))
                             break
                         elif data_cell == int(scouting_data['Setup']['Match']):
                             # Update an existing row
                             team_sheet = update_data_row( team_sheet, attr_row, data_row, scouting_data )
                             shutil.copyfile(input_dir+verified_file, input_dir+verified_file.replace('verified','processed'))
                             break
                             
                         # Jump over the next two rows
                         i += 2
                # ######################################################### #
                         
            xlsx_workbook.save(output_file)
                             

def update_data_row( team_sheet, attr_row, data_row, scouting_data ):
    max_column= team_sheet.max_column
    
    # start with the setup data fields, which are at the start of the row
    scouting_section = scouting_data['Setup']
    # iterate over all columns
    for i in range(1,max_column+1):
        attr_value = team_sheet.cell(row=attr_row,column=i).value
        # if the row of attributes has an empty cell, that marks the end of the
        # list of scouting attributes, so break out of the loop
        if attr_value is None:
            break
        attr_value = attr_value.replace(' ','_')
        
        data_value = scouting_section.get(attr_value)
        if scouting_data.get(attr_value) is not None:
            scouting_section = scouting_data[attr_value]
        else:
            try:
                team_sheet.cell(row=data_row, column=i).value = int(data_value)
            except:
                team_sheet.cell(row=data_row, column=i).value = data_value

    return team_sheet

'''
def copy_range( sheet, start_col, start_row, end_col, end_row ):
    range_selected = []
    #Loops through selected Rows
    for i in range(start_row,end_row + 1,1):
        #Appends the row to a RowSelected list
        row_selected = []
        for j in range(start_col,end_col+1,1):
            row_selected.append(sheet.cell(row = i, column = j))
        #Adds the RowSelected List and nests inside the rangeSelected
        range_selected.append(row_selected)
 
    return range_selected
    
def paste_range( dest_sheet, start_col, start_row, end_col, end_row, copied_cells ):
    #Paste data from copyRange into template sheet
    count_row = 0
    for i in range(start_row,end_row+1,1):
        count_col = 0
        for j in range(start_col,end_col+1,1):
            
            dest_sheet.cell(row = i, column = j).value = copiedData[count_row][count_col]
            count_col += 1
        count_row += 1    
'''
    
def create_team_sheet(xlsx_workbook, team_name):
    blank_team_sheet = xlsx_workbook.get_sheet_by_name('Team Sheet')
    decision_sheet = xlsx_workbook.get_sheet_by_name('Decision Sheet')

    # find an opening in the decision sheet for this team's data
    dest_cell = '%s1' % get_column_letter( decision_sheet.max_column+1)
        
    team_sheet = xlsx_workbook.copy_worksheet(blank_team_sheet)
    team_sheet.title = team_name
    team_sheet['A1'] = team_name
    
    # Define start Range(target_start) in the new Worksheet
    min_col, min_row, max_col, max_row = range_boundaries(dest_cell)

    source_range = 'D1:E30'
    
    for row, row_cells in enumerate(decision_sheet[source_range], min_row):
        for column, cell in enumerate(row_cells, min_col):
            cell_value = cell.value
            if cell_value is not None:
                cell_value = cell_value.replace('Team 1', team_name)
                cell_value = cell_value.replace('Team Sheet', team_name)
                
                if cell_value.startswith('=PRODUCT'):
                    frags = cell_value.split(',')
                    frags[1] = frags[1].replace('D', get_column_letter(column-1))
                    cell_value = ','.join(frags)
                    
                if cell_value.startswith('=SUM'):
                    cell_value = cell_value.replace('E', get_column_letter(column) )
            
            decision_sheet.cell(row=row, column=column).value = cell_value
            decision_sheet.cell(row=row, column=column).fill = copy(cell.fill)
            
    return team_sheet

if __name__ == "__main__":

    print 'Processing files...'

    # command line options handling
    parser = OptionParser()
    
    parser.add_option(
        "-l","--processloop",dest="processloop", default='0',
        help='Process Team Files')
    parser.add_option(    
        "-a","--aliases",dest="comp_alias_file",default='ScoutingAppEventAliases.txt',
        help="Competition Alias Configuration File")
    
    # Parse the command line arguments
    (options,args) = parser.parse_args()

    global_config = {}
    
    ConfigUtils.read_config(global_config, './config/ScoutingAppConfig.txt')

    logger = Logger.get_logger('./config', 'logging.conf', 'scouting.fileproc')
    global_config['logger'] = logger

    # load the competition alias file if one is specified
    if options.comp_alias_file != '':
        comp_alias_file = './config/' + options.comp_alias_file        
        logger.debug('Loading Competition Alias file: %s' % comp_alias_file)
        CompAlias.read_comp_alias_config(comp_alias_file)

    session = DbSession.open_db_session((global_config['db_name'] + global_config['this_season']), DataModel)

    counter = 0
    done = False

    while not done:
        try:
            global_config['logger'].debug( 'Scanning for new files to process' )
            print 'Scanning for new files to process'
            start_time = datetime.datetime.now()
            
            competition = global_config['this_competition'] + global_config['this_season']
            competition_dir = './static/data/' + competition
            input_dir = competition_dir + '/ScoutingData/'
    
            # the following section replaces the original file processing. The new model
            # reads in the JSON scouting data files and writes the information to a specially 
            # formatted spreadsheet output file. Over time, we may migrate back to a more
            # traditional model like before, but for 2019, this is what we're going to do.
            output_file = input_dir + '2019 ScoutingSystem %s.xlsx' % global_config['this_competition']
            process_json_files(global_config, competition, output_file, input_dir)
            process_issue_files(global_config, input_dir)
            process_debrief_files(global_config, input_dir)
            global_config['logger'].debug( 'Scan complete, elapsed time - %s' % (str(datetime.datetime.now()-start_time)) )
            print 'Scan complete, elapsed time - %s' % (str(datetime.datetime.now()-start_time))

            '''
            attrdef_filename = WebCommonUtils.get_attrdef_filename(short_comp=global_config['this_competition'])
            if attrdef_filename is None:
                global_config['logger'].debug( 'No Attribute Definitions, Skipping Process Files' )
                print 'No Attribute Definitions, Skipping Process Files'
            else:
                attr_definitions = AttributeDefinitions.AttrDefinitions(global_config)
                attr_definitions.parse(attrdef_filename)
                    
                process_files(global_config, attr_definitions, input_dir)
                process_issue_files(global_config, input_dir)
                process_debrief_files(global_config, input_dir)
                global_config['logger'].debug( 'Scan complete, elapsed time - %s' % (str(datetime.datetime.now()-start_time)) )
                print 'Scan complete, elapsed time - %s' % (str(datetime.datetime.now()-start_time))
            '''
        except Exception, e:
            global_config['logger'].debug('Exception Caught Processing Files: %s' % str(e) )
            print 'Exception Caught Processing Files: %s' % str(e)
            traceback.print_exc(file=sys.stdout)
            exc_type, exc_value, exc_traceback = sys.exc_info()
            exception_info = traceback.format_exception(exc_type, exc_value,exc_traceback)
            for line in exception_info:
                line = line.replace('\n','')
                global_config['logger'].debug(line)
                print line
        
        if int(options.processloop) > 0:
            time.sleep(int(options.processloop))
        else:
            done = True

    session.remove()
    
    print 'Process files complete...'




    
